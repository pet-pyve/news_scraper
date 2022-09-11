#imports
from urllib.request import Request, urlopen
from bs4 import BeautifulSoup
import re
import datetime
import openpyxl
import os

links = [] #a list which contains urls to iterate through
num = 1
page_string = "page/" + str(num) +"/" #the date of article which makes part of url


#strip the formatting of the text in the article to be parsed
def strip_formatting(string):
    string = string.lower()
    string = string.replace("\n", " ")
    string = string.replace("'", "")
    string = re.sub(r"([.!?,'/()])", r" \1 ", string)

    return string

while len(links) < 2000: #only grab 2000 urls to search through

    news_page = "https://www.naturalnews.com/all-posts/" + page_string  #crafting the url
    hdr = {'User-Agent': 'Mozilla/5.0'}
    req = Request(news_page,headers=hdr)
    #open the url
    page = urlopen(req)

    #decode the html of the page
    soup = BeautifulSoup(page, "html.parser")
    articles = soup.findAll("div",  attrs={"class": "f-p-title"})


    for div in articles: #all the link elements on the page which have 'news', please take them
        a = div.findAll("a", href = True)
        for url in a:
            links.append("https://www.naturalnews.com/" + str(url['href']))
    print(len(links))

    num += 1
    page_string = "page/" + str(num) +"/"


excel_document = openpyxl.load_workbook('training_data.xlsx') #open the excel document to store article in
ws = excel_document.worksheets[0] #get the right sheet
row = 10651 #start from row 2
number_for_print = 0

for tasty_link in links: #for every link in the list
    try:
        column = 1 #start at column one
        news_page = tasty_link
        hdr = {'User-Agent': 'Mozilla/5.0'}
        req = Request(news_page,headers=hdr)
        page = urlopen(req)
        soup = BeautifulSoup(page, "html.parser")

        #finding the text in the article
        name_box = soup.findAll("div", attrs={"class": "entry-content"})
        text = []

        for i in name_box:
            text.append(strip_formatting(i.text.strip()))

        ws.cell(row=row, column=column).value = "natural news" #from the abc
        column += 1

        ws.cell(row=row, column=column).value = tasty_link #record the url

        column += 1
        ws.cell(row=row, column=column).value = str(text) #article text
        column += 1

        ws.cell(row=row, column=column).value = 1 #the article classifier
        column += 1
        row += 1
        number_for_print += 1
        print(str(number_for_print) + " out of " + str(len(links)))

    except Exception:
        print("oh no")
        continue

excel_document.save('training_data.xlsx') #remember to sve
