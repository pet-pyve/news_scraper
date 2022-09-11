#imports
import re
import datetime
import openpyxl
import os

excel_document = openpyxl.load_workbook('training_data.xlsx') #open the excel document to store article in
ws = excel_document.worksheets[0] #get the right sheet
ws_new = excel_document.worksheets[1]
body = []
row_num = 1
number_for_print = 1

for row in ws.values:
    column = 1

    print(str(number_for_print))
    number_for_print += 1

    if str(row[2]) not in body and str(row[2]) != "[]":

        ws_new.cell(row=row_num, column=column).value = str(row[0]) #from the abc
        column += 1

        ws_new.cell(row=row_num, column=column).value = str(row[1]) #record the url
        column += 1

        ws_new.cell(row=row_num, column=column).value = str(row[2]) #article text
        column += 1

        ws_new.cell(row=row_num, column=column).value = str(row[3]) #the article classifier
        column += 1
        row_num += 1

        body.append(str(row[2]))


excel_document.save('training_data.xlsx') #remember to sve
