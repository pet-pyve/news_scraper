#imports
import urllib.request
from bs4 import BeautifulSoup
import re
import datetime
import openpyxl
import os

tasty_urls = [] #a list which contains urls to iterate through
date_string = "2018-08-27" #the date of article which makes part of url
page_one = "&page=1" #the page number of archives

#strip the formatting of the text in the article to be parsed
def strip_formatting(string):
    string = string.lower()
    string = string.replace("\n", " ")
    string = string.replace("'", "")
    string = re.sub(r"([.!?,'/()])", r" \1 ", string)

    return string

while len(tasty_urls) < 2000: #only grab 2000 urls to search through

    #crafting the url
    news_page = "http://www.abc.net.au/news/archive/?date=" + str(date_string) + page_one

    #open the url
    page = urllib.request.urlopen(news_page)

    #decode the html of the page
    soup = BeautifulSoup(page, "html.parser")

    for a in soup.find_all('a', href=True): #all the link elements on the page which have 'news', please take them
        if ("/news/" + date_string) in a['href']:
            if ("http://www.abc.net.au" + a['href']) not in tasty_urls: #only if you havent already collected the url
                tasty_urls.append("http://www.abc.net.au" + a['href'])

    date_string = datetime.datetime.strptime(date_string, "%Y-%m-%d") - datetime.timedelta(1) #decrease the date by one day
    date_string = str(date_string)[:10] #formatting the date


excel_document = openpyxl.load_workbook('training_data.xlsx') #open the excel document to store article in
ws = excel_document.worksheets[0] #get the right sheet
row = 2 #start from row 2

for tasty_link in tasty_urls: #for every link in the list
    column = 1 #start at column one
    news_page = tasty_link
    page = urllib.request.urlopen(news_page) #go to the page
    soup = BeautifulSoup(page, "html.parser")

    #finding the text in the article
    name_box = soup.find("div", attrs={"class": "article section"})
    paragraph = name_box.find_all("p")

    text = []

    for i in paragraph:
        try:
            this_is_bad = i["class"] #ignore all html with a specific class
        except KeyError:
            text.append(strip_formatting(i.text.strip())) # taking only raw article text


    ws.cell(row=row, column=column).value = "ABC" #from the abc
    column += 1

    ws.cell(row=row, column=column).value = tasty_link #record the url

    column += 1
    ws.cell(row=row, column=column).value = str(text) #article text
    column += 1

    ws.cell(row=row, column=column).value = 0 #the article classifier
    column += 1
    row += 1


excel_document.save('training_data.xlsx') #remember to sve
