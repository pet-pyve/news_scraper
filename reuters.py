
import urllib.request
from bs4 import BeautifulSoup
import re
import datetime
import openpyxl
import os

tasty_urls = []
date_string = "20180825"


def strip_formatting(string):
    string = string.lower()
    string = string.replace("\n", " ")
    string = string.replace("'", "")
    string = re.sub(r"([.!?,'/()])", r" \1 ", string)

    return string

while len(tasty_urls) < 2000:

    news_page = "https://www.reuters.com/resources/archive/us/" + str(date_string) + ".html"

    page = urllib.request.urlopen(news_page)

    soup = BeautifulSoup(page, "html.parser")

    for a in soup.find_all('a'):
        if "article" in a['href']:
            if a['href'] not in tasty_urls:
                tasty_urls.append(a['href'])


    date_string = datetime.datetime.strptime(date_string, "%Y%m%d") - datetime.timedelta(1)
    date_string = str(date_string)[:4] + str(date_string)[5:7] + str(date_string)[8:10]

    print(len(tasty_urls))


excel_document = openpyxl.load_workbook('training_data.xlsx')
ws = excel_document.worksheets[0]
row = 2007

for tasty_link in tasty_urls:
    column = 1
    news_page = tasty_link
    try:
        page = urllib.request.urlopen(news_page)
    except:
        continue

    soup = BeautifulSoup(page, "html.parser")

    name_box = soup.find("div", attrs={"class": "StandardArticleBody_body"})
    if name_box != None:
        paragraph = name_box.find_all("p")

        text = []

        for i in paragraph:
            try:
                this_is_bad = i["class"]
            except KeyError:
                text.append(strip_formatting(i.text.strip()))

        ws.cell(row=row, column=column).value = "reuters"
        column += 1
        print(tasty_link)
        ws.cell(row=row, column=column).value = tasty_link

        column += 1
        ws.cell(row=row, column=column).value = str(text)
        column += 1
        ws.cell(row=row, column=column).value = 0
        column += 1
        row += 1
        print(row)

excel_document.save('training_data.xlsx')
