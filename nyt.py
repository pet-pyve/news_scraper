#imports
import urllib.request
from bs4 import BeautifulSoup
import re
import datetime
import openpyxl
import os

links = [] #a list which contains urls to iterate through
date_string = "20181217" #the date of article which makes part of url

#strip the formatting of the text in the article to be parsed
def strip_formatting(string):
    string = string.lower()
    string = string.replace("\n", " ")
    string = string.replace("'", "")
    string = re.sub(r"([.!?,'/()])", r" \1 ", string)

    return string

while len(links) < 2000: #only grab 2000 urls to search through
    num = 0
    #crafting the url

    news_page = "https://spiderbites.nytimes.com/2018/articles_2018_12_0000" + str(num) + ".html"

    #open the url
    page = urllib.request.urlopen(news_page)

    #decode the html of the page
    soup = BeautifulSoup(page, "html.parser")
    articles = soup.findAll("a", href = True)

    print("artircle", len(articles))

    for heading in articles: #all the link elements on the page which have 'news', please take them
        if "https://www.nytimes.com/2018" in str(heading):
            print(heading)
            links.append(heading['href'])
                #print(i['href'])


    num += 1
print("done")
excel_document = openpyxl.load_workbook('training_data.xlsx') #open the excel document to store article in
ws = excel_document.worksheets[0] #get the right sheet
row = 4038 #start from row 2

for tasty_link in links: #for every link in the list
    column = 1 #start at column one
    news_page = tasty_link
    page = urllib.request.urlopen(news_page) #go to the page
    soup = BeautifulSoup(page, "html.parser")

        #finding the text in the article
    name_box = soup.find(attrs={"class": "css-1572rug"})
    text = []
    try:
        for i in name_box:
            text.append(strip_formatting(i.text.strip()))
    except Exception:
        pass


    ws.cell(row=row, column=column).value = "nyt" #from the abc
    column += 1

    ws.cell(row=row, column=column).value = tasty_link #record the url

    column += 1
    ws.cell(row=row, column=column).value = str(text) #article text
    column += 1

    ws.cell(row=row, column=column).value = 0 #the article classifier
    column += 1
    row += 1
    print(row)


excel_document.save('training_data.xlsx') #remember to sve
